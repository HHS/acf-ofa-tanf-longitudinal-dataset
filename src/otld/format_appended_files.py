"""Format appended files

This module formats appended TANF financial data.
"""

import os

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import numbers
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from otld import combine_appended_files


def format_pd_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Format pandas data frame columns.

    Apply transformations to column values.

    Args:
        df (pd.DataFrame): Data frame to format.

    Returns:
        pd.DataFrame: Formatted data frame.
    """
    numeric = df.select_dtypes(include=[np.number]).columns
    df[numeric] = df[numeric].map(lambda x: f"{x:,.0f}")

    return df


def add_table(ws: Worksheet, displayName: str, ref: str):
    """Add an Excel table to a worksheet.

    Args:
        ws (Worksheet): Worksheet to add table to.
        displayName (str): Name of the table.
        ref (str): Range of cells to convert to a table.
    """
    tab = Table(displayName=displayName, ref=ref)

    style = TableStyleInfo(
        name="TableStyleLight8", showRowStripes=True, showColumnStripes=True
    )
    tab.tableStyleInfo = style

    ws.add_table(tab)


def format_openpyxl_worksheet(ws: Worksheet):
    """Format openpyxl worksheet.

    Args:
        ws (Worksheet): Worksheet to format.
    """
    # Adjust columns
    for column in range(ws.max_column):
        column += 1
        column_letter = get_column_letter(column)
        column = ws.column_dimensions[column_letter]
        # Adjust width
        column.width = 25.0

    # Align right
    for i, row in enumerate(ws.rows):
        if i == 0:
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            continue

        for cell in row[2:]:
            cell.alignment = Alignment(horizontal="right")

            # Check if the value is a number and, if so, format as currency.
            # No need to worry about years being formatted as currencies since they
            # appear in the first 2 columns of each row and therefore are excluded.
            try:
                float(cell.value)
                cell.number_format = numbers.FORMAT_CURRENCY_USD
            except ValueError:
                pass


def export_workbook(frames: dict, path: str):
    """Export a dictionary of data frames as an Excel Workbook.

    Args:
        frames (dict): A dictionary of pandas DataFrames.
        path (str): The path at which to output the Excel workbook.
    """
    # Load csv into workbook
    # Adapted from https://stackoverflow.com/questions/12976378/openpyxl-convert-csv-to-excel
    wb = openpyxl.Workbook()
    ws = wb.active

    i = 0
    for frame in frames:
        if i == 0:
            ws = wb.active
            ws.title = frame
        else:
            wb.create_sheet(frame)
            ws = wb[frame]

        i += 1

        df = frames[frame]
        df = dataframe_to_rows(df.reset_index(), index=False)

        for row in df:
            ws.append(row)

        add_table(ws, frame, ws.dimensions)
        format_openpyxl_worksheet(ws)

    # Export
    wb.save(path)


def main():
    """Entry point for script to format appended files"""

    frames = combine_appended_files.main()
    export_workbook(frames, os.path.join(out_dir, "FinancialDataWide.xlsx"))

    frames["FinancialData"] = []
    for frame in frames:
        if frame == "FinancialData":
            continue
        frames[frame] = frames[frame].melt(
            var_name="Category", value_name="Amount", ignore_index=False
        )
        frames[frame]["Level"] = frame
        frames["FinancialData"].append(frames[frame])

    frames["FinancialData"] = pd.concat(frames["FinancialData"])
    for frame in list(frames.keys()):
        if frame != "FinancialData":
            del frames[frame]

    export_workbook(frames, os.path.join(out_dir, "FinancialDataLong.xlsx"))


if __name__ == "__main__":
    from otld.paths import out_dir

    main()
