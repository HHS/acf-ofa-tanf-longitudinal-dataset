"""Class to handle tracking the sources of line numbers"""

import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from otld.utils.openpyxl_utils import set_column_widths


class LineTracker:
    """Class to handle tracking the sources of line numbers"""

    def __init__(self):
        """Initiate sources dictionary"""
        self.sources = {}

    def export(self, path: str):
        """Export sources dictionary to an Excel workbook

        Args:
            path (str): Path to an existing Excel workbook/Excel workbook to be created.
        """
        if not os.path.exists(path):
            sheet_name = list(self.sources.keys())[0]
            sheet_name = str(sheet_name)
            pd.DataFrame().to_excel(path, sheet_name=sheet_name)

        workbook = load_workbook(path)

        for year in self.sources:
            year_str = str(year)

            df = (
                pd.DataFrame()
                .from_dict(self.sources[year], orient="columns")
                .explode(["BaseColumns", "RenamedColumns"])
            )

            if year_str in workbook.sheetnames:
                del workbook[year_str]

            # Place sheet at the correct location
            if year == 1997:
                index = 0
            else:
                sheets = [int(name) for name in workbook.sheetnames]
                nearest = year
                while nearest not in sheets:
                    nearest -= 1
                index = sheets.index(nearest) + 1

            workbook.create_sheet(year_str, index=index)
            worksheet = workbook[year_str]

            # Adapted from https://stackoverflow.com/questions/17326973/is-there-a-way-to-auto-adjust-excel-column-widths-with-pandas-excelwriter
            column_widths = [
                max(df[column].astype(str).map(len).max(), len(column)) for column in df
            ]

            # Add data to worksheet and set column widths
            df = dataframe_to_rows(df, index=False)

            for row in df:
                worksheet.append(row)

            worksheet = set_column_widths(worksheet, column_widths)

        workbook.save(path)
