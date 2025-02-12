"""Common openpyxl utilities"""

__all__ = [
    "delete_empty_columns",
    "get_merged_value",
    "get_column_names",
    "export_workbook",
    "long_notes",
]

import openpyxl
from openpyxl.styles import numbers
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


# Appropriated from Gemini code sample
def delete_empty_columns(
    worksheet: Worksheet,
) -> Worksheet:
    """Delete columns with all NA/empty values

    Args:
        worksheet (Worksheet): Worksheet to delete columns from.

    Returns:
        Worksheet: Worksheet with empty columns removed
    """
    # Iterate over columns in reverse order to avoid index issues when deleting
    for column in range(worksheet.max_column, 0, -1):
        is_empty = True
        for row in worksheet.iter_rows(min_col=column, max_col=column):
            cell = row[0]
            if cell.value is not None:
                is_empty = False
                break

        if is_empty:
            worksheet.delete_cols(column)


def get_merged_value(worksheet: Worksheet, cell: openpyxl.cell.cell.Cell) -> str:
    """Return the value in the first cell of a range of merged cells

    Args:
        worksheet (Worksheet): The worksheet to search in for merged cells.
        cell (openpyxl.cell.cell.Cell): The cell to check for in the merged cells.

    Returns:
        str: The value of the first merged cell or the cell's value if it is not in a
        range of merged cells.
    """
    merged_cells = list(worksheet.merged_cells)
    for merged_cell in merged_cells:
        if cell.coordinate in merged_cell:
            first_cell = merged_cell.coord.split(":")[0]
            return worksheet[first_cell].value

    return cell.value if cell.value else ""


def get_column_names(worksheet: Worksheet) -> list[str]:
    """Get the column names of an Excel worksheet

    This function iterates through the rows in an Excel worksheet until all columns
    have a non-empty value and then returns these as candidate column names.

    Args:
        worksheet (Worksheet): An Excel worksheet.

    Returns:
        list[str]: A list of potential column names
    """
    i = 0
    columns = []
    # If not all column names are present, concatenate current row with next
    # row
    for row in worksheet.rows:
        i += 1
        if i == 1:
            continue

        if columns:
            columns = [
                (column + " " + get_merged_value(worksheet, row[j])).strip()
                for j, column in enumerate(columns)
            ]
        else:
            columns = [get_merged_value(worksheet, row[j]) for j in range(len(row))]

        if all(columns):
            break

    columns = [column.strip() for column in columns]
    columns = [column for column in columns if column]
    return columns, i


def set_column_widths(worksheet: Worksheet, column_width: int | list[int]) -> Worksheet:
    """Set the column widths for an Excel worksheet.

    Args:
        worksheet (Worksheet): An openpyxl worksheet.
        column_width (int | list[int]): Integer, or list of integers specifying column
        width(s)

    Returns:
        Worksheet: Updated worksheet
    """
    for index in range(worksheet.max_column):
        width = column_width if isinstance(column_width, int) else column_width[index]
        index += 1
        column_letter = get_column_letter(index)
        dimensions = worksheet.column_dimensions[column_letter]
        # Adjust width
        dimensions.width = width

    return worksheet


def add_table(ws: Worksheet, displayName: str, ref: str):
    """Add an Excel table to a worksheet.

    Args:
        ws (Worksheet): Worksheet to add table to.
        displayName (str): Name of the table.
        ref (str): Range of cells to convert to a table.
    """
    tab = Table(displayName=displayName, ref=ref)

    style = TableStyleInfo(
        name="TableStyleLight8", showRowStripes=True, showColumnStripes=True
    )
    tab.tableStyleInfo = style

    ws.add_table(tab)


def format_openpyxl_worksheet(
    ws: Worksheet, skip_cols: int = 2, number_format: str = numbers.FORMAT_CURRENCY_USD
):
    """Format openpyxl worksheet.

    Args:
        ws (Worksheet): Worksheet to format.
    """
    # Adjust columns
    for column in range(ws.max_column):
        column += 1
        column_letter = get_column_letter(column)
        column = ws.column_dimensions[column_letter]
        # Adjust width
        column.width = 25.0

    # Align right
    for i, row in enumerate(ws.rows):
        if i == 0:
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            continue

        for cell in row[skip_cols:]:
            cell.alignment = Alignment(horizontal="right")

            # Check if the value is a number and, if so, format as currency.
            # No need to worry about years being formatted as currencies since they
            # appear in the first 2 columns of each row and therefore are excluded.
            try:
                float(cell.value)
                cell.number_format = number_format
            except ValueError:
                pass


def export_workbook(
    frames: dict,
    path: str,
    drop: list[str] = [],
    format_options: dict = {},
    footnotes: dict[list[list]] = {},
):
    """Export a dictionary of data frames as an Excel Workbook.

    Args:
        frames (dict): A dictionary of pandas DataFrames.
        path (str): The path at which to output the Excel workbook.
        drop (list[str], optional): List of columns to drop from the data frames. Defaults to [].
    """
    # Load csv into workbook
    # Adapted from https://stackoverflow.com/questions/12976378/openpyxl-convert-csv-to-excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # If there are footnotes, then the keys should be a subset of frames.keys()
    if footnotes:
        frame_set = set(frames.keys())
        note_set = set(footnotes.keys())
        assert frame_set.issuperset(
            note_set
        ), f"All keys in footnotes ({footnotes.keys()}) should be in frames ({frames.keys()})"

    for i, frame in enumerate(frames):
        if i == 0:
            ws = wb.active
            ws.title = frame
        else:
            wb.create_sheet(frame)
            ws = wb[frame]

        df = frames[frame].copy()
        if drop:
            df.drop(drop, inplace=True, axis=1)

        df = dataframe_to_rows(df.reset_index(), index=False)

        for row in df:
            ws.append(row)

        add_table(ws, frame, ws.dimensions)

        # If there are footnotes, add them after the table. Otherwise continue.
        if footnotes:
            notes = footnotes.get(frame, [])

            for row in notes:
                row = row.copy()
                while len(row) < ws.max_column:
                    row += [""]

                ws.append(row)

        format_openpyxl_worksheet(ws, **format_options)

    # Export
    wb.save(path)


def long_notes(new_key: str, footnotes: dict[list[list]]):
    footnotes = footnotes.copy()
    footnotes[new_key] = [note for notes in footnotes.values() for note in notes]

    for key in list(footnotes.keys()):
        if key == new_key:
            continue
        del footnotes[key]

    for key, value in footnotes.items():
        notes = []
        for note in value:
            if note not in notes:
                notes.append(note)

        footnotes[key] = notes

    return footnotes
